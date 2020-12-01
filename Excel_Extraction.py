import pandas as pd
import os
import re
from fuzzywuzzy import fuzz
from operator import itemgetter, add
from openpyxl import worksheet
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import collections

full_path = os.path.abspath(os.getcwd())
# file_name = r"C:\Users\willlee\Desktop\Manufacturing Issue\2020\Test Station Part Number.xlsx"
red_background = PatternFill(bgColor=colors.BLUE)
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ["istext($AA2)"]  # <---need to change to dynamic


def text_splitter(text, regex_expression, take_which_group=1):
    matches = re.finditer(regex_expression, text, re.MULTILINE)

    for matchNum, match in enumerate(matches, start=1):

        print("Match {matchNum} was found at {start}-{end}: {match}".format(matchNum=matchNum, start=match.start(),
                                                                            end=match.end(), match=match.group()))

        for groupNum in range(0, len(match.groups())):
            groupNum = groupNum + 1

            print("Group {groupNum} found at {start}-{end}: {group}".format(groupNum=groupNum,
                                                                            start=match.start(groupNum),
                                                                            end=match.end(groupNum),
                                                                            group=match.group(groupNum)))
    #   if take_which_group == 1:
    #       return match.group(1)
    #   elif take_which_group == 2:
    #       return match.group(2)
    #   else:
    #       print ("No Defined Group. Assume Group 1")
    #       return match.group(1)
    try:
        match.group(take_which_group)
    except NameError:
        match = None
    if match == None:
        return "No Defined"
    else:
        return match.group(take_which_group)


def each_Location(value_items, df_in, list_columns=['Station Name', 'TE PN']):
    # Create dictonary
    dictonary = {}
    two_d_array = []
    # Convert the list
    list_value = list(value_items)
    # Get the dataframe that relate to the BU
    df_interest = df_in.loc[list_value, list_columns]
    df_PN = df_interest.groupby(['TE PN'])
    # Get keys
    key = df_PN.groups.keys()
    values = df_PN.groups.values()
    for keys_value in df_PN.groups:
        list_to_store_station_name = []
        print(keys_value)
        value_list = list(df_PN.groups.get(keys_value).values)
        for index_value in value_list:
            list_to_store_station_name.append(df_in.loc[index_value, 'Station Name'])
        # Assign into the dictonary
        dictonary[keys_value] = list_to_store_station_name
    two_d_array.append(dictonary)
    #        df = df_interest.loc[list_v,'Station Name']
    return two_d_array


#        each_Location(list_value, df_interest, ['Station Name'])

def filter_data_type_check(list_array, some_text):
    ratio_marks = []
    station_id_list = []
    always_same = True
    already_inside = False
    different_list = []
    # first level checking
    for dictionary_detail in list_array:
        ratio_marks.append(fuzz.partial_ratio(dictionary_detail[0], some_text[0]))
    # look for the all max marks
    m = max(ratio_marks)
    first_level_check_index = [i for i, j in enumerate(ratio_marks) if j == m]
    # filter the array that have max marks
    filter_two = itemgetter(*first_level_check_index)(list_array)
    # use to check the first column is the same
    station_id_list = [i[0] for i in filter_two]
    print(station_id_list)
    print([item for item, count in collections.Counter(station_id_list).items() if count >= 1])
    different_list = [item for item, count in collections.Counter(station_id_list).items() if count >= 1]
    # for i in range (0, len(filter_two)):
    #     if i != len(filter_two) -1:
    #         if filter_two[i][0] == filter_two[i+1][0]:
    #             pass #goto next element
    #         else:
    #             always_same = False
    #             #store first value
    #             if len(different_list) == 0:
    #                 different_list.append(filter_two[i][0])
    #                 different_list.append(filter_two[i+1][0])
    #             else:
    #                 #Checking it already written
    #                 for list_in_different_list in different_list:
    #                     if filter_two[i][0] == list_in_different_list or  filter_two[i+1][0] == list_in_different_list :
    #                         already_inside = True
    #                         break
    #
    #                     else:
    #                         already_inside = False
    #                 if not already_inside:
    #                     different_list = filter_two[i][0]
    #     else:
    #         pass
    # #store the last result
    # print (different_list)
    # different_list.append(filter_two[len(filter_two)-1][0])
    return different_list


def filter_data_id_check(list_array, some_text):
    filter_two = []
    ratio_marks = []
    for dictionary_detail in list_array:
        ratio_marks.append(fuzz.ratio(dictionary_detail[1], some_text[1]))
    print(ratio_marks)
    m = max(ratio_marks)
    first_level_check_index = [i for i, j in enumerate(ratio_marks) if j == m]
    # filter the array that have max marks
    for index in first_level_check_index:
        print(list_array[index])
        str_list = list_array[index][0] + list_array[index][1] + list_array[index][2]
        filter_two.append(str_list)
    #    filter_two = itemgetter(*first_level_check_index)(list_array)
    #    for i in range(0, len(filter_two)):
    #        pass
    print(filter_two)
    return filter_two


def find_nearest_station_rule_two(list_array, some_text):
    filter_element = []
    first_level_check_index = []
    ratio_marks = []
    two_d_list = []
    list_correction_value = []
    filter_station_type = filter_data_type_check(list_array, some_text)  # Filter  the station type
    # get all same station type
    for i, j in enumerate(list_array):
        if j[0] == filter_station_type[0]:
            str_station = j[0] + j[1] + j[2]
            first_level_check_index.append(str_station)
    #    first_level_check_index = [j for i, j in enumerate(list_array) if j[0] == filter_station_type]
    filter_station_id = filter_data_id_check(list_array, some_text)  # Filter  the station id

    list_filter_two = []

    for i in first_level_check_index:
        list_filter_two.append(i)
    for i in filter_station_id:
        list_filter_two.append(i)
    print(list_filter_two)
    ratio_marks_total = []
    whole_station = []
    ratio_marks_station_id = []  # Clear the list
    ratio_marks_station_type = []  # Clear the list
    # append back to become a list

    # Once adding do the correction again Checking Ratio
    for dictionary_detail in list_filter_two:
        ratio_marks_station_type.append(fuzz.ratio(dictionary_detail, some_text))
    # ratio_marks_total =  list(map(add, ratio_marks_station_id, ratio_marks_station_type))
    m = max(ratio_marks_station_type)
    # get the index
    two_level_check_index = [i for i, j in enumerate(ratio_marks_station_type) if j == m]
    multiple_string_match = ""
    for j in two_level_check_index:
        multiple_string_match = multiple_string_match + list_filter_two[j] + ";"
        list_correction_value.append(list_filter_two[j])

    print(multiple_string_match)
    # Checking if there is multiple value
    a = [item for item, count in collections.Counter(list_correction_value).items() if count >= 1]
    print(a)
    return a


# ([A-Z]{0,4})\s?([0-9]+\w+)?\s?(PEN|PXIE|PXI)?
# ([A-Z]{0,4})\s?([0-9]+)\s?(PEN|PXIE|PXI)?
# ([A-Z]{0,4})\s?(?:([0-9|A-Z]+))\s?(PEN|PXIE|PXI)
#  ori value = ([A-Z]{0,4})\s?([0-9]+)\s?(PEN|PXIE|PXI)?
def find_nearest_station_rule_one(df, regex=r'([A-Z]{0,4})\s?(?:([0-9|A-Z]+))\s?(PEN|PXIE|PXI|NIH)'):
    #    regex = r'([A-Z]{3,4})\s?(\w+)\s?(PXI|PXIE)?'
    one_d_array = []
    if isinstance(df, pd.DataFrame):
        two_d_array = []
        for single_station_name in df['Station Name']:
            station_id = text_splitter(single_station_name, regex, 1)
            station_number = text_splitter(single_station_name, regex, 2)
            location = text_splitter(single_station_name, regex, 3)
            # one_d_array.append([station_id,station_number,location])
            two_d_array.append([station_id, station_number, location])
            one_d_array = []
        return two_d_array
    elif isinstance(df, str):
        station_id = text_splitter(df, regex, 1)
        station_number = text_splitter(df, regex, 2)
        location = text_splitter(df, regex, 3)
        return [station_id, station_number, location]
    else:
        return


def find_nearest_station_rule_three(list_array, some_text):
    ratio_marks = []
    station_list = []
    # first level checking
    for dictionary_detail in list_array:
        ratio_marks.append(fuzz.ratio(dictionary_detail[1], some_text[1]))
    m = max(ratio_marks)
    first_level_check_index = [i for i, j in enumerate(ratio_marks) if j == m]
    # filter the array that have max marks
    filter_two = itemgetter(*first_level_check_index)(list_array)
    for single_element in filter_two:
        station_list.append(single_element[0] + single_element[1] + single_element[2])
    return station_list


def find_nearest_station_rule_all(df, text):
    list_dictionary = find_nearest_station_rule_one(df)
    # find_nearest_station_rule_one(df,regex= )
    # Process with different rule?
    single_text = find_nearest_station_rule_one(text)
    # Data Checking to check the existance of the content
    if single_text[0]:  # station id exists
        return find_nearest_station_rule_two(list_dictionary, single_text)
    else:  # no station id
        return find_nearest_station_rule_three(list_dictionary, single_text)


def correction_station_name(station_name,dictionary_file):
    file_name = dictionary_file
    #file_name = r"C:\Users\willlee\Desktop\Manufacturing Issue\2020\Test Station Part Number.xlsx"
    xl = pd.ExcelFile(file_name)
    print(xl.sheet_names)
    xl.close()
    # 2. Open Excel File Read with the Correct Sheet
    df = pd.read_excel(io=file_name, sheet_name=xl.sheet_names[2])
    a = find_nearest_station_rule_all(df, station_name)  # df is a dictionary
    return a


def correct_temp_file(file_name):
    str_text = ""
    str_location = ""
    str_TEPN = ""
    # 1. Get Every Sheet Name
    xl = pd.ExcelFile(file_name)
    print(xl.sheet_names)
    xl.close()
    # 2. Open Excel File Read with the Correct Sheet
    df = pd.read_excel(io=file_name, sheet_name=xl.sheet_names[2])
    print(df.head(5))
    dictonary = {}

    # Checking whether it is in undefined
    xl = pd.ExcelFile(r"C:\Users\willlee\Desktop\ABCD.xlsx")
    print(xl.sheet_names)
    xl.close()
    workbook = load_workbook(filename=r"C:\Users\willlee\Desktop\ABCD.xlsx")
    # looking for the correct sheet
    for s in range(len(workbook.sheetnames)):
        if workbook.sheetnames[s] == xl.sheet_names[0]:
            workbook.active = s  # set which file is active
            break
    worksheet = workbook.active
    #    worksheet = workbook[xl.sheet_names[0]]
    # apply conditional rule
    worksheet.conditional_formatting.add("A2:AA500", rule)
    #    a = find_nearest_station_rule_all(df, r"1302PEN")
    # 2. Open Excel File Read with the Correct Sheet
    df_abcd = pd.read_excel(io=r"C:\Users\willlee\Desktop\ABCD.xlsx",
                            sheet_name=xl.sheet_names[0])  # <---need to change to dynamic
    df_filter = df_abcd.loc[df_abcd['Station Name'] == 'Undefined']
    if len(df_filter) != 0:
        worksheet['AA1'] = 'Nearest Correction Value'  # <---need to change to dynamic
        # worksheet['AB1'] = 'Nearest Correction Location'  # <---need to change to dynamic
        # worksheet['AC1'] = 'Nearest Correction PN'  # <---need to change to dynamic
    for index, row in df_filter.iterrows():
        print(row)
        station_name = row['Test Station(s)']
        a = find_nearest_station_rule_all(df, station_name)
        # looking for column A - index should be A6
        i = 2
        while True:
            str_row = "A" + str(i)
            if worksheet[str_row].value == index:
                break
            elif worksheet[str_row].value == None:
                print("Row empty")
                break
            else:
                i = i + 1
            if i >= 1000:
                print("Not found anything")
                break
        column_station = "AA" + str(i)  # <---need to change to dynamic
        # column_location = "AB" + str(i)  # <---need to change to dynamic
        # column_TEPN = "AC" + str(i)  # <---need to change to dynamic
        # Concate to the string
        for station_naming in a:
            # # retrieve TE PN
            # try:
            #     location = list(df.loc[df['Station Name'] == station_naming]['Location'])[0]
            # except:
            #     location = '-'
            # try:
            #     TePN = list(df.loc[df['Station Name'] == station_naming]['TE PN'])[0]
            # except:
            #     TePN = "-"
            str_text = str_text + station_naming + ";"
            # str_location = str_location + location + ";"
            # str_TEPN = str_TEPN + TePN + ";"
        worksheet[column_station] = str_text
        # worksheet[column_location] = str_location
        # worksheet[column_TEPN] = str_TEPN
        str_text = ""
        # str_location = ""
        # str_TEPN = ""
    workbook.save(filename=r"C:\Users\willlee\Desktop\ABCD_1.xlsx")
    workbook.close()


def main(file_name):
    # ! need to define the column to write the suggestion value
    dictonary = {}
    xl = pd.ExcelFile(file_name)
    print(xl.sheet_names)
    xl.close()
    # 2. Open Excel File Read with the Correct Sheet
    df = pd.read_excel(io=file_name, sheet_name=xl.sheet_names[2])
    print(df.head(5))

    # 3. Need to read first row for every column to use in group by
    # df.columns = df.columns.str.strip()
    # print (df.columns)
    # 4. Get the planner Code and Module level]
    df_grouping = df.groupby(['Location'])
    print(df_grouping)
    Dict = {}
    elements = []
    # elements.append([])
    for key in list(df_grouping.groups.items()):
        two_d_array = each_Location(key[1], df)
        #        print (two_d_array)
        dictonary[key[0]] = two_d_array
    print(dictonary)
    return dictonary


#        for i in value:
#            print (df.loc[i])
#            print (type(df.loc[i]))
# df.loc[i][0]
#            elements = part_number_process(elements,df,i)

#        Dict[key] = elements
#        del elements
#        elements = []

# Iterate over key/value for debug purposes # Comment out if need to run faster
#    for key,value in Dict.items():
#        print (key,':',value)

if __name__ == "__main__":
    main(r"C:\Users\willlee\Desktop\Manufacturing Issue\2020\Test Station Part Number.xlsx")
#     correct_temp_file(r"C:\Users\willlee\Desktop\Manufacturing Issue\2020\Test Station Part Number.xlsx")
